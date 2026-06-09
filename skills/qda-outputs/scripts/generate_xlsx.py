#!/usr/bin/env python3
"""
generate_xlsx.py
Generates the QDA coded corpus Excel workbook from v3 coded speech files.

Usage (run from project root):
    python skills/qda-outputs/scripts/generate_xlsx.py

Output:
    analysis/results/qda_v3_coded_corpus.xlsx
"""

import re
import sys
from pathlib import Path
from collections import OrderedDict

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not found. Run: pip install openpyxl --break-system-packages")
    sys.exit(1)

# ── Configuration ─────────────────────────────────────────────────────────────

PROJECT_ROOT = Path(__file__).parent.parent.parent.parent  # skills/qda-outputs/scripts/ → project root
CODED_DIR    = PROJECT_ROOT / "analysis" / "coded_speeches" / "v3"
OUTPUT_PATH  = PROJECT_ROOT / "analysis" / "results" / "qda_v3_coded_corpus.xlsx"

# Speech display names and metadata (order = chronological)
SPEECH_ORDER = [
    "Debate Presidencial 2023",
    "WEF Davos 2024",
    "Apertura Sesiones Ordinarias 2024 (142°)",
    "UN General Assembly 79th",
    "WEF Davos 2025",
    "Apertura Sesiones Ordinarias 2025 (143°)",
    "UN General Assembly 80th",
    "WEF Davos 2026",
    "Apertura Sesiones Ordinarias 2026 (144°)",
]

# Map frontmatter dates → speech display names
DATE_TO_SPEECH = {
    "2023-10-08": ("Debate Presidencial 2023", "Buenos Aires", "Debate"),
    "2024-01-17": ("WEF Davos 2024", "Davos", "Speech"),
    "2024-03-01": ("Apertura Sesiones Ordinarias 2024 (142°)", "Buenos Aires", "Speech"),
    "2024-09-24": ("UN General Assembly 79th", "New York", "Speech"),
    "2025-01-23": ("WEF Davos 2025", "Davos", "Speech"),
    "2025-03-01": ("Apertura Sesiones Ordinarias 2025 (143°)", "Buenos Aires", "Speech"),
    "2025-09-24": ("UN General Assembly 80th", "New York", "Speech"),
    "2026-01-22": ("WEF Davos 2026", "Davos", "Speech"),
    "2026-03-01": ("Apertura Sesiones Ordinarias 2026 (144°)", "Buenos Aires", "Speech"),
}

# Sub-code normalized display names (underscore_key → "Display Name")
SUBCODE_DISPLAY = {
    "Anti_Casta": "Anti-Casta (Elite Frame)",
    "Anti_State_Anti_Bureaucratic": "Anti-State / Anti-Bureaucratic",
    "Delegitimization_of_Regulation": "Delegitimization of Regulation",
    "Deregulation_Benefits": "Deregulation Benefits",
    "Market_Primacy": "Market Primacy",
    "Rejection_of_Multilateral_Governance": "Rejection of Multilateral Governance",
    "Sustainability_as_Ideological_or_Harmful": "Sustainability as Ideological or Harmful",
    "Attack_on_Public_Research_Institutions": "Attack on Public Research Institutions",
    "Attack_on_Universities": "Attack on Universities",
    "Attack_on_Experts_and_Technocrats": "Attack on Experts & Technocrats",
    "Climate_Skepticism_or_Dismissal": "Climate Skepticism or Dismissal",
    "Pseudo_Science_or_Ideological_Science": "Pseudo-Science / Ideological Science",
    "Rejection_of_Expert_Mediation": "Rejection of Expert Mediation",
    "Civilizational_Defense": "Civilizational Defense",
    "Neo_Marxism": "Neo-Marxism",
    "Woke_Bundling": "Woke Bundling",
    "Reported_Denial": "Reported Denial",
    "Selective_Empiricism": "Selective Empiricism",
    "Temporal_Appropriation": "Temporal Appropriation",
}

DIM_FULL = {
    "D1.1": "D1.1 — State Intervention & Governance",
    "D1.2": "D1.2 — Public Science & Expertise",
    "D1.3": "D1.3 — Cultural Backlash",
    "IND":  "Inductive",
}


# ── Parser ────────────────────────────────────────────────────────────────────

def parse_frontmatter_date(text):
    m = re.search(r'^date:\s*(\d{4}-\d{2}-\d{2})', text, re.MULTILINE)
    return m.group(1) if m else None


def parse_code_tag(tag_str):
    """Parse 'D1.1 > Sub_code' or 'INDUCTIVE > Sub_code | ...' → (dim, sub_key)."""
    m = re.match(r'([A-Z0-9\.]+)\s*>\s*([A-Za-z_]+)', tag_str.strip())
    if not m:
        return None, None
    dim_raw, sub_raw = m.group(1), m.group(2)
    dim = "IND" if dim_raw == "INDUCTIVE" else dim_raw
    return dim, sub_raw


def clean_passage(text):
    """Strip markdown role labels and formatting from passage text."""
    text = re.sub(r'\*\*\[[^\]]+\]\*\*:\s*', '', text)
    text = re.sub(r'\*\*([^*]+)\*\*', r'\1', text)
    text = re.sub(r'\*([^*]+)\*', r'\1', text)
    return ' '.join(text.split()).strip()


def parse_coded_file(path):
    """
    Read a v3 coded speech file and return a list of:
        {'passage': str, 'dim': str, 'sub_key': str, 'memo': str,
         'speech': str, 'date': str, 'venue_type': str}
    """
    raw = path.read_text(encoding='utf-8')

    # Remove YAML frontmatter
    body = re.sub(r'^---\n.*?---\n', '', raw, count=1, flags=re.DOTALL)
    # Remove coding summary section
    body = re.sub(r'\n---\n## Coding Summary.*$', '', body, flags=re.DOTALL)
    body = re.sub(r'\n## Coding Summary.*$', '', body, flags=re.DOTALL)

    date = parse_frontmatter_date(raw) or "0000-00-00"
    meta = DATE_TO_SPEECH.get(date)
    if meta is None:
        # Try to infer from filename (YYYY-MM-DD-*)
        m = re.match(r'(\d{4}-\d{2}-\d{2})', path.name)
        date = m.group(1) if m else date
        meta = DATE_TO_SPEECH.get(date, (path.stem, "Unknown", "speech"))

    speech_name, venue, stype = meta
    venue_type = f"{venue} · {stype.title()}"

    lines = body.split('\n')
    entries = []
    last_passage = ""
    text_buffer = []

    def flush_buffer():
        nonlocal text_buffer, last_passage
        if text_buffer:
            candidate = clean_passage(' '.join(text_buffer))
            if len(candidate) > 15:
                last_passage = candidate
            text_buffer = []

    for line in lines:
        stripped = line.strip()
        # Code annotation line
        m = re.match(r'`\[CODE:\s*([^\]]+)\]`\s*[—–\-]+\s*\*?(?:Memo:)?\s*(.*)', stripped)
        if m:
            if text_buffer:
                flush_buffer()
            tag = m.group(1)
            memo = m.group(2).strip().strip('*').strip()
            dim, sub_key = parse_code_tag(tag)
            if dim and sub_key and last_passage:
                entries.append({
                    'passage': last_passage,
                    'dim': dim,
                    'sub_key': sub_key,
                    'memo': memo,
                    'speech': speech_name,
                    'date': date,
                    'venue_type': venue_type,
                })
            continue

        # Skip separators, headers, context notes
        if not stripped or stripped.startswith('---') or stripped.startswith('#'):
            flush_buffer()
            continue
        if re.match(r'^\*\[', stripped) or re.match(r'^\[Not coded', stripped):
            flush_buffer()
            continue

        # Regular text — add to buffer; if we just flushed a code block, start fresh
        cleaned = re.sub(r'\*\*([^*]+)\*\*', r'\1', stripped)
        cleaned = re.sub(r'\*([^*]+)\*', r'\1', cleaned).strip()
        if cleaned and len(cleaned) > 5:
            if entries and entries[-1].get('_flushed'):
                text_buffer = [cleaned]
            else:
                text_buffer.append(cleaned)

    flush_buffer()
    return entries


# ── Excel builder ─────────────────────────────────────────────────────────────

# Style helpers
def header_fill(r, g, b):
    return PatternFill("solid", fgColor=f"{r:02X}{g:02X}{b:02X}")

def thin_border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

TITLE_FONT  = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
HEADER_FONT = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
DATA_FONT   = Font(name='Calibri', size=10)
WRAP_ALIGN  = Alignment(wrap_text=True, vertical='top')
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')

DIM_COLORS = {
    'D1.1': ('1E40AF', 'FFFFFF'),  # blue, white text
    'D1.2': ('B45309', 'FFFFFF'),  # amber
    'D1.3': ('991B1B', 'FFFFFF'),  # red
    'IND':  ('374151', 'FFFFFF'),  # gray
    'TOTAL':('1F2937', 'FFFFFF'),
}


def build_workbook(entries, speech_list):
    wb = openpyxl.Workbook()

    # ── Sheet 1: Coded Passages ────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Coded Passages'

    # Title row
    ws.merge_cells('A1:J1')
    ws['A1'] = "Milei Climate Narratives — v3 Coded Passages  ·  Data source for all summary sheets"
    ws['A1'].font = Font(name='Calibri', bold=True, size=12, color='111827')
    ws['A1'].fill = PatternFill('solid', fgColor='E0F2FE')
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 22

    # Headers
    headers = ['#', 'Dimension', 'Sub-dimension', 'Dimension (Full)', 'Sub-code (Full)',
               'Quoted Passage', 'Analytical Memo', 'Speech', 'Date', 'Venue / Type']
    col_widths = [5, 8, 30, 34, 34, 55, 60, 25, 12, 18]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(2, ci, h)
        cell.font = HEADER_FONT
        cell.fill = PatternFill('solid', fgColor='1E3A5F')
        cell.alignment = CENTER_ALIGN
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 16

    # Data rows
    for i, e in enumerate(entries, 1):
        r = i + 2
        sub_display = SUBCODE_DISPLAY.get(e['sub_key'], e['sub_key'].replace('_', ' '))
        dim_full = DIM_FULL.get(e['dim'], e['dim'])
        row_vals = [
            i, e['dim'], sub_display, dim_full, sub_display,
            e['passage'], e['memo'], e['speech'], e['date'], e['venue_type']
        ]
        bg = '1E40AF' if e['dim'] == 'D1.1' else \
             'B45309' if e['dim'] == 'D1.2' else \
             '991B1B' if e['dim'] == 'D1.3' else '374151'
        for ci, v in enumerate(row_vals, 1):
            cell = ws.cell(r, ci, v)
            cell.font = DATA_FONT
            cell.alignment = WRAP_ALIGN
            if ci <= 2:
                cell.fill = PatternFill('solid', fgColor=bg + '22')
        ws.row_dimensions[r].height = 60

    # Freeze header rows
    ws.freeze_panes = 'A3'

    # ── Sheet 2: Freq — By Dimension ──────────────────────────────────────────
    wd = wb.create_sheet('Freq — By Dimension')
    wd.merge_cells('A1:L1')
    wd['A1'] = "Code Frequency by Dimension × Speech  —  values update automatically from 'Coded Passages'"
    wd['A1'].font = Font(bold=True, size=11)
    wd['A1'].fill = PatternFill('solid', fgColor='E0F2FE')

    sp_cols = {s: get_column_letter(i + 2) for i, s in enumerate(speech_list)}
    header2 = ['Dimension', 'Dimension (Full)'] + speech_list + ['TOTAL', '', 'Dimension', 'Total']
    for ci, h in enumerate(header2, 1):
        cell = wd.cell(2, ci, h)
        cell.font = HEADER_FONT
        cell.fill = PatternFill('solid', fgColor='1E3A5F')
        cell.alignment = CENTER_ALIGN
    wd.column_dimensions['A'].width = 8
    wd.column_dimensions['B'].width = 36
    for col in sp_cols.values():
        wd.column_dimensions[col].width = 22
    total_col = get_column_letter(len(speech_list) + 3)
    wd.column_dimensions[total_col].width = 10

    dims_order = ['D1.1', 'D1.2', 'D1.3', 'IND']
    data_rows = {}
    for dr, dim in enumerate(dims_order, 3):
        wd.cell(dr, 1, dim).font = Font(bold=True)
        wd.cell(dr, 2, DIM_FULL.get(dim, dim))
        for sp, col in sp_cols.items():
            ci = ord(col) - ord('A') + 1
            wd.cell(dr, ci).value = (
                f"=COUNTIFS('Coded Passages'!$B:$B,$A{dr},'Coded Passages'!$H:$H,{col}$2)"
            )
        sum_col = len(speech_list) + 3
        sum_letter = get_column_letter(sum_col)
        wd.cell(dr, sum_col).value = f"=SUM(C{dr}:{get_column_letter(sum_col-1)}{dr})"
        # Mini-table
        wd.cell(dr, sum_col + 2, dim)
        wd.cell(dr, sum_col + 3).value = f"=I{dr}"
        data_rows[dim] = dr

    # TOTAL row
    tr = len(dims_order) + 3
    wd.cell(tr, 1, 'TOTAL').font = Font(bold=True)
    wd.cell(tr, 2, 'All Dimensions').font = Font(bold=True)
    first_dr, last_dr = 3, 3 + len(dims_order) - 1
    for ci in range(3, len(speech_list) + 4):
        col = get_column_letter(ci)
        wd.cell(tr, ci).value = f"=SUM({col}{first_dr}:{col}{last_dr})"
        wd.cell(tr, ci).font = Font(bold=True)

    # ── Sheet 3: Freq — By Sub-code ───────────────────────────────────────────
    ws2 = wb.create_sheet('Freq — By Sub-code')
    ws2.merge_cells('A1:L1')
    ws2['A1'] = "Code Frequency by Sub-code × Speech  —  values update automatically from 'Coded Passages'"
    ws2['A1'].font = Font(bold=True, size=11)
    ws2['A1'].fill = PatternFill('solid', fgColor='E0F2FE')

    for ci, h in enumerate((['Dim', 'Sub-code'] + speech_list + ['TOTAL', '', 'Sub-code', 'Total']), 1):
        cell = ws2.cell(2, ci, h)
        cell.font = HEADER_FONT
        cell.fill = PatternFill('solid', fgColor='1E3A5F')
        cell.alignment = CENTER_ALIGN
    ws2.column_dimensions['A'].width = 7
    ws2.column_dimensions['B'].width = 38
    for col in sp_cols.values():
        ws2.column_dimensions[col].width = 20

    # Collect ordered sub-codes
    subcodes_by_dim = OrderedDict()
    for e in entries:
        dim = e['dim']
        sub = SUBCODE_DISPLAY.get(e['sub_key'], e['sub_key'].replace('_', ' '))
        subcodes_by_dim.setdefault(dim, OrderedDict())[sub] = True
    # Add all known sub-codes even if not yet in corpus
    for dim, subs in [
        ('D1.1', ['Anti-Casta (Elite Frame)', 'Anti-State / Anti-Bureaucratic',
                  'Delegitimization of Regulation', 'Deregulation Benefits', 'Market Primacy',
                  'Rejection of Multilateral Governance', 'Sustainability as Ideological or Harmful']),
        ('D1.2', ['Attack on Public Research Institutions', 'Attack on Universities',
                  'Attack on Experts & Technocrats', 'Climate Skepticism or Dismissal',
                  'Pseudo-Science / Ideological Science', 'Rejection of Expert Mediation']),
        ('D1.3', ['Civilizational Defense', 'Neo-Marxism', 'Woke Bundling']),
        ('IND',  ['Reported Denial', 'Selective Empiricism', 'Temporal Appropriation']),
    ]:
        subcodes_by_dim.setdefault(dim, OrderedDict())
        for s in subs:
            subcodes_by_dim[dim].setdefault(s, True)

    sc_rows = {}
    sr = 3
    for dim in ['D1.1', 'D1.2', 'D1.3', 'IND']:
        for sub in subcodes_by_dim.get(dim, {}).keys():
            ws2.cell(sr, 1, dim)
            ws2.cell(sr, 2, sub)
            for sp, col in sp_cols.items():
                ci = ord(col) - ord('A') + 1
                ws2.cell(sr, ci).value = (
                    f"=COUNTIFS('Coded Passages'!$B:$B,$A{sr},'Coded Passages'!$C:$C,$B{sr},'Coded Passages'!$H:$H,{col}$2)"
                )
            sum_col = len(speech_list) + 3
            ws2.cell(sr, sum_col).value = f"=SUM(C{sr}:{get_column_letter(sum_col-1)}{sr})"
            ws2.cell(sr, sum_col + 2, f"{dim}·{sub[:18]}")
            ws2.cell(sr, sum_col + 3).value = f"=I{sr}"
            sc_rows[(dim, sub)] = sr
            sr += 1

    # TOTAL row
    ws2.cell(sr, 1, 'TOTAL').font = Font(bold=True)
    ws2.cell(sr, 2, 'All Sub-codes').font = Font(bold=True)
    for ci in range(3, len(speech_list) + 4):
        col = get_column_letter(ci)
        ws2.cell(sr, ci).value = f"=SUM({col}3:{col}{sr-1})"
        ws2.cell(sr, ci).font = Font(bold=True)

    # ── Sheet 4: Speech Overview ───────────────────────────────────────────────
    ws3 = wb.create_sheet('Speech Overview')
    ws3.merge_cells('A1:I1')
    ws3['A1'] = "Speech-Level Overview  —  D1 / Inductive counts update automatically from 'Coded Passages'  (D2 = 0 in v3)"
    ws3['A1'].font = Font(bold=True, size=11)
    ws3['A1'].fill = PatternFill('solid', fgColor='E0F2FE')

    ov_headers = ['Speech', 'Date', 'Venue', 'Type', 'D1 Codes', 'D2 Codes', 'Inductive', 'Total', 'Notes']
    for ci, h in enumerate(ov_headers, 1):
        cell = ws3.cell(2, ci, h)
        cell.font = HEADER_FONT
        cell.fill = PatternFill('solid', fgColor='1E3A5F')
    ws3.column_dimensions['A'].width = 28
    ws3.column_dimensions['B'].width = 12
    ws3.column_dimensions['C'].width = 14
    ws3.column_dimensions['D'].width = 10

    # Use speech metadata
    speech_meta = {v[0]: (k, v[1], v[2]) for k, v in DATE_TO_SPEECH.items()}
    for si, sp in enumerate(speech_list, 3):
        date, venue, stype = speech_meta.get(sp, ("", "", ""))
        ws3.cell(si, 1, sp)
        ws3.cell(si, 2, date)
        ws3.cell(si, 3, venue)
        ws3.cell(si, 4, stype)
        ws3.cell(si, 5).value = f"=COUNTIFS('Coded Passages'!$H:$H,$A{si},'Coded Passages'!$B:$B,\"D1*\")"
        ws3.cell(si, 6).value = f"=COUNTIFS('Coded Passages'!$H:$H,$A{si},'Coded Passages'!$B:$B,\"D2*\")"
        ws3.cell(si, 7).value = f"=COUNTIFS('Coded Passages'!$H:$H,$A{si},'Coded Passages'!$B:$B,\"IND\")"
        ws3.cell(si, 8).value = f"=E{si}+F{si}+G{si}"
        ws3.cell(si, 9, "D2=0 in v3 (speech corpus)")

    tr = len(speech_list) + 3
    ws3.cell(tr, 1, 'TOTAL').font = Font(bold=True)
    for ci in range(5, 9):
        col = get_column_letter(ci)
        ws3.cell(tr, ci).value = f"=SUM({col}3:{col}{tr-1})"
        ws3.cell(tr, ci).font = Font(bold=True)

    return wb


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    if not CODED_DIR.exists():
        print(f"ERROR: Coded speeches directory not found: {CODED_DIR}")
        sys.exit(1)

    coded_files = sorted(CODED_DIR.glob("*.md"))
    coded_files = [f for f in coded_files if not f.name.startswith("frequency_tally")]
    if not coded_files:
        print(f"ERROR: No coded speech .md files found in {CODED_DIR}")
        sys.exit(1)

    print(f"Found {len(coded_files)} coded speech files:")

    all_entries = []
    for path in coded_files:
        entries = parse_coded_file(path)
        print(f"  {path.name}: {len(entries)} code annotations")
        all_entries.extend(entries)

    print(f"\nTotal code annotations: {len(all_entries)}")

    # Determine which speeches are present and their order
    present_speeches = []
    for sp in SPEECH_ORDER:
        if any(e['speech'] == sp for e in all_entries):
            present_speeches.append(sp)
    # Add any speeches not in the predefined order
    for e in all_entries:
        if e['speech'] not in present_speeches:
            present_speeches.append(e['speech'])

    print(f"Speeches in workbook: {present_speeches}")

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = build_workbook(all_entries, present_speeches)
    wb.save(OUTPUT_PATH)
    print(f"\n✓ Saved: {OUTPUT_PATH}")
    print(f"  Sheets: {wb.sheetnames}")
    print(f"  Coded Passages rows: {len(all_entries)}")


if __name__ == '__main__':
    main()

ame__ == '__main__':
    main()
